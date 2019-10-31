import openpyxl
import sys
import re
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from copy import copy

tab_val = [[] for x in range(5)]
tab_val2 = [[] for x in range(5)]
interface_policy_group = []
dealt_with = []
my_tuple = ()
list_of_same_destination_port = []
list_of_same_leaf_id = []
list_of_list_of_same = []
wb1 = openpyxl.load_workbook("/home/nico/scripts/ACI_Implementation_Plan_Paris_draft_Copy.xlsx", data_only=True)
ws1 = wb1["interface_selector"]
wb2 = openpyxl.load_workbook("interface_selector_access.xlsx", data_only=True)
ws2 = wb2.active
for row in ws2.iter_rows(max_col=ws2.max_column, min_row=2, max_row=1161):
    if row[0].value == "vpc":
        tab_val2[0].append(row[0].value)
        tab_val2[1].append(row[1].value)
        tab_val2[2].append(row[2].value)
        tab_val2[3].append(row[3].value)
        tab_val2[4].append(row[4].value)
    else:
        tab_val[0].append(row[0].value)
        tab_val[1].append(row[1].value)
        tab_val[2].append(row[2].value)
        tab_val[3].append(row[3].value)
        tab_val[4].append(row[4].value)

for i in tab_val2[4]:
    if i not in dealt_with:
        interface_policy_group.append(i)
        dealt_with.append(i)

for i in interface_policy_group:
    for index, j in enumerate(tab_val2[4]):
        if j == i:
            list_of_same_destination_port.append(tab_val2[3][index])
            list_of_same_leaf_id.append(tab_val2[2][index])
    my_tuple = (copy(list_of_same_destination_port), copy(list_of_same_leaf_id))
    list_of_list_of_same.append(copy(my_tuple))
    list_of_same_destination_port.clear()
    list_of_same_leaf_id.clear()

rows = []
values = []

for row in ws1.iter_rows(max_col=ws1.max_column, min_row=2, max_row=2):
    for cell in row:
        values.append(cell.value)

for a, b, c, d, e in zip(*tab_val):
        values[0] = "INT_1_" + str(d)
        values[2] = "LEF_" + str(c) + "_PROD_TN_IPR"
        values[4] = str(d)
        values[6] = str(d)
        values[8] = str(e)
        rows.append(copy(values))

for index, i in enumerate(list_of_list_of_same):
    if i[0].count(i[0][0]) == len(i[0]):
        values[0] = "INT_1_" + str(i[0][0])
    else:
        values[0] = "INT_1_" + str(min(i[0])) + "_TO_1_" + str(max(i[0]))
    values[2] = "LEF_" + str(min(i[1])) + "_" + str(max(i[1])) + "_PROD_TN_IPR"
    values[4] = str(min(i[0]))
    values[6] = str(max(i[0]))
    values[8] = interface_policy_group[index]
    values[9] = "accbundle"
    rows.append(copy(values))

for row in range(2, len(rows) + 2):
    for col in range(1, len(values) + 1):
        ws1.cell(column=col, row=row, value=rows[row - 2][col - 1])

wb1.save(filename="ACI_Implementation_Plan_Paris_draft_Copy.xlsx")