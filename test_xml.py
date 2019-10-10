import xml.etree.ElementTree as ET
from xml.etree.ElementTree import ElementTree
import openpyxl
import sys
import re
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

wb = openpyxl.load_workbook("/home/nico/scripts/ACI_Implementation_Plan_Paris_draft_Copy.xlsx", data_only=True)
ws = wb["Fabric Configuration Steps"]

list_of_step = []
list_of_step_tuple = []

first = re.escape("{{")
second = re.escape("}}")

def get_interface_policy_group_type(row, ws):
    return ws["D" + str(row + 1)].value

def parse_excel(sheet, row, column):
    for col in sheet.iter_cols(min_row=1, max_col=sheet.max_column, max_row=1):
        for cell in col:
            if cell.value == column:
                coordinate = cell.coordinate
                xy = coordinate_from_string(coordinate)
                row_value = xy[1] + row
                return sheet[xy[0] + str(row_value)].value

def get_config_steps(sheet, list_of_step):
    for col in sheet.iter_cols(min_col=3, max_col=3, min_row=1, max_row=sheet.max_row):
        for cell in col:
            list_of_step.append(cell.value)
    return [x for x in list_of_step if x]

def get_tuple(sheet, list_of_step):
    for step in list_of_step:
        for col in sheet.iter_cols(min_col=3, max_col=3, min_row=2, max_row=sheet.max_row):
            for cell in col:
                if cell.value == step:
                    xml_file = cell.offset(column=1).value
                    mode = cell.offset(column=2).value
                    list_of_step_tuple.append((step, xml_file, mode))
                    break

def main_work(ws, step, secondroot, tree, filename):
    if ws.title == "interface_policy_group":
        handle_interface_group(secondroot, ws, tree, filename)
    else:
        for rows_index in range(1, ws.max_row):
            subtree = ET.parse(step[1])
            thirdroot = subtree.getroot()
            secondroot.append(thirdroot)
            replace_variable(thirdroot, rows_index)
            tree.write(filename, encoding="UTF-8", xml_declaration=True)

def replace_variable(root, rows_index):
    for t in root.iter():
        for i, j in t.attrib.items():
            if j == "":
                continue
            list_of_matches = re.findall((first + "\w+" + second), j)
            list_of_matches = [x for x in list_of_matches if x]
            for y in list_of_matches:
                list_match = list(y)
                for index in range(-2, 1):
                    del(list_match[index])
                del(list_match[0])
                list_match = "".join(list_match)
                j = j.replace(y, str(parse_excel(ws, rows_index, list_match)))
            t.attrib[i] = j

def handle_interface_group(secondroot, ws, tree, filename):
        thirdroot = ET.Element("infraFuncP")
        secondroot.insert(0, thirdroot)
        for rows_index in range(1, ws.max_row):
            interface_type = get_interface_policy_group_type(rows_index, ws)
            if interface_type == "vPC":
                subtree = ET.parse("/home/nico/scripts/infraAccBndlGrp_vpc.xml")
                fourthroot = subtree.getroot()
                thirdroot.append(fourthroot)
                replace_variable(fourthroot, rows_index)
            elif interface_type == "Access":
                subtree = ET.parse("/home/nico/scripts/infraAccBndlGrp_access.xml")
                fourthroot = subtree.getroot()
                thirdroot.append(fourthroot)
                replace_variable(fourthroot, rows_index)
            elif interface_type == "pc":
                subtree = ET.parse("/home/nico/scripts/infraAccBndlGrp_pc.xml")
                fourthroot = subtree.getroot()
                thirdroot.append(fourthroot)
                replace_variable(fourthroot, rows_index)
            tree.write(filename, encoding="UTF-8", xml_declaration=True)

list_of_step = get_config_steps(ws, list_of_step)
ws = wb["build_tasks"]
get_tuple(ws, list_of_step)
root_infra = ET.Element("polUni")
second_root_infra = ET.Element("infraInfra")
root_infra.insert(0, second_root_infra)
tree_infra = ET.ElementTree(root_infra)
root_tenant = ET.Element("polUni")
second_root_tenant = ET.Element("fv Tenant")
root_tenant.insert(0, second_root_tenant)
tree_tenant = ET.ElementTree(root_tenant)
root_blank = ET.Element("polUni")
tree_blank = ET.ElementTree(root_blank)
for step in list_of_step_tuple:
    for sheet in wb:
        if sheet.title == step[0]:
            ws = wb[step[0]]
            if step[2] == "Fabric External Access Policy":
                main_work(ws, step, second_root_infra, tree_infra, "file_infra")
            elif step[2] == "Tenant Base Configuration":
                main_work(ws, step, second_root_tenant, tree_tenant, "file_tenant")
            else:
                main_work(ws, step, root_blank, tree_blank, "file_blank")
            break