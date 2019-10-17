import openpyxl
import sys
import re
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from copy import copy

mode = []
description = []
leaf_id = []
destination_port = []
interface_policy_group = []
wb1 = openpyxl.load_workbook("/home/nico/scripts/ACI_Implementation_Plan_Paris_draft_Copy.xlsx", data_only=True)
ws1 = wb1["interface_selector"]
wb2 = openpyxl.load_workbook("interface_selector_access.xlsx", data_only=True)
ws2 = wb2.active
for row in ws2.iter_rows(max_col=ws2.max_column, min_row=2, max_row=564):
    for cell in row:
        if cell.col_idx == 1:
            description.append(cell.value)
        elif cell.col_idx == 2:
            mode.append(cell.value)
        elif cell.col_idx == 3:
            leaf_id.append(cell.value)
        elif cell.col_idx == 4:
            destination_port.append(cell.value)
        elif cell.col_idx == 5:
            interface_policy_group.append(cell.value)

rows = []
values = []

for row in ws1.iter_rows(max_col=ws1.max_column, min_row=2, max_row=2):
    for cell in row:
        values.append(cell.value)

for a, b, c, d, e in zip(description, mode, leaf_id, destination_port, interface_policy_group):
    values[0] = "INT_1_" + str(d)
    values[2] = "LEF_" + str(c) + "_PROD_TN_IPR"
    values[4] = str(d)
    values[6] = str(d)
    values[8] = str(e)
    rows.append(copy(values))

for row in range(2, len(description) + 2):
    for col in range(1, len(values) + 1):
        ws1.cell(column=col, row=row, value=rows[row - 2][col - 1])
wb1.save(filename="ACI_Implementation_Plan_Paris_draft_Copy.xlsx")